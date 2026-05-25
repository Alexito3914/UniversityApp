"""
Importa titulaciones, asignaturas, profesores y ofertas desde HORARIOS_25_26.xlsx (Listado).
Opcionalmente carga la cuadrícula de referencia del decanato (--apply-schedule).
"""
import re
import unicodedata
from datetime import time
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.core.exceptions import ValidationError
from django.db import transaction

from horarios.models import (
    AcademicYear,
    Classroom,
    Course,
    DegreeProgram,
    Professor,
    ProfessorAvailability,
    Schedule,
    ScheduleEntry,
    Subject,
    SubjectOffering,
    TimeSlotConfig,
)

STUDY_MAP = {
    'Gininf': ('INF', 'Grado en Ingeniería Informática'),
    'GRobotica': ('ROB', 'Grado en Ingeniería Robótica'),
    'GTransporte': ('TEL', 'Grado en Ingeniería Telemática'),
}

DOUBLE_DEGREE = ('DINFROB', 'Doble Grado Informática + Robótica')
EPS_DEGREE_CODES = ('INF', 'ROB', 'TEL', 'DINFROB')

DEGREE_CLASSROOMS = {
    'INF': ('LAB-INF', 'Laboratorio Informática', 30),
    'ROB': ('LAB-ROB', 'Laboratorio Robótica', 24),
    'TEL': ('LAB-TEL', 'Laboratorio Telemática', 28),
}

DAY_MAP = {
    'LUNES': 'MON',
    'MARTES': 'TUE',
    'MIÉRCOLES': 'WED',
    'MIERCOLES': 'WED',
    'JUEVES': 'THU',
    'VIERNES': 'FRI',
}

SCHEDULE_GROUP_MAP = {
    '1INF': ('INF', 1, 'Grupo A'),
    '2INF': ('INF', 2, 'Grupo A'),
    '3INF': ('INF', 3, 'Grupo A'),
    '4INF-TI': ('INF', 4, '4INF-TI'),
    '4INF-IS': ('INF', 4, '4INF-IS'),
    '1ROB': ('ROB', 1, 'Grupo 1'),
    '2ROB': ('ROB', 2, 'Grupo 2'),
    '3ROB': ('ROB', 3, 'Grupo 3'),
    '4ROB': ('ROB', 4, 'Grupo 4'),
}

TIME_BLOCKS = [
    (time(9, 0), time(11, 0)),
    (time(11, 0), time(13, 0)),
    (time(13, 0), time(15, 0)),
    (time(15, 30), time(17, 30)),
    (time(17, 30), time(19, 30)),
    (time(19, 30), time(21, 30)),
]


def _ascii_upper(text):
    if not text:
        return ''
    normalized = unicodedata.normalize('NFKD', str(text))
    return normalized.encode('ascii', 'ignore').decode('ascii').upper()


def _prof_code(full_name):
    """Genera código tipo RMONTUFAR a partir de «Apellidos, Nombre»."""
    if not full_name:
        return ''
    parts = str(full_name).split(',')
    if len(parts) < 2:
        return _ascii_upper(full_name).replace(' ', '')[:12]
    last = parts[0].strip().split()[0] if parts[0].strip() else ''
    first = parts[1].strip().split()[0] if parts[1].strip() else ''
    return (_ascii_upper(first[:1]) + _ascii_upper(last)) if last else ''


def _parse_course_number(raw):
    if not raw:
        return None
    m = re.search(r'(\d+)', str(raw))
    return int(m.group(1)) if m else None


def _course_shift(degree_code, course_num):
    if degree_code == 'DINFROB':
        return 'AFTERNOON' if course_num >= 4 else 'MORNING'
    return 'AFTERNOON' if course_num == 4 else 'MORNING'


def _group_name(study, course_num, doble_val):
    if study == 'Gininf' and course_num == 4:
        if doble_val == 4:
            return '4INF-TI'
        if doble_val == 5:
            return '4INF-IS'
    if study == 'GRobotica' and doble_val not in (None, True, False, ''):
        try:
            return f'Grupo {int(doble_val)}'
        except (TypeError, ValueError):
            pass
    if doble_val in (None, True, False, ''):
        return 'Grupo A'
    try:
        return f'Grupo {int(doble_val)}'
    except (TypeError, ValueError):
        return 'Grupo A'


def _weekly_sessions(h1, h2):
    count = 0
    for h in (h1, h2):
        if h and str(h).strip() not in ('', 'True', 'False', 'None'):
            count += 1
    return count if count else 2


def _parse_professor_name(raw):
    raw = str(raw or '').strip()
    if not raw:
        return 'Sin', 'Asignar', 'sin.asignar@ucjc.local'
    if ',' in raw:
        last_part, first_part = raw.split(',', 1)
        last_name = last_part.strip()
        first_name = first_part.strip().split()[0] if first_part.strip() else 'Docente'
    else:
        bits = raw.split()
        first_name = bits[0] if bits else 'Docente'
        last_name = ' '.join(bits[1:]) if len(bits) > 1 else bits[0]
    slug = _ascii_upper(last_name).replace(' ', '').replace(',', '')[:20].lower()
    email = f'{slug or "prof"}.{_ascii_upper(first_name[:1]).lower()}@ucjc.local'
    return first_name, last_name, email


DEFAULT_EXCEL = Path(__file__).resolve().parents[3] / 'data' / 'HORARIOS_25_26.xlsx'


class Command(BaseCommand):
    help = 'Importa asignaturas/profesores/ofertas desde HORARIOS_25_26.xlsx (hoja Listado).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=str(DEFAULT_EXCEL),
            help='Ruta al Excel del decanato',
        )
        parser.add_argument('--year', default='2026-2027', help='Nombre del año académico (debe coincidir con el horario borrador)')
        parser.add_argument(
            '--apply-schedule',
            action='store_true',
            help='Carga la cuadrícula de las hojas HORARIOS 1º/2º CUATRIMESTRE',
        )
        parser.add_argument(
            '--clear-eps',
            action='store_true',
            help='Borra ofertas/sesiones EPS (INF/ROB/TEL/DINFROB) antes de importar',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'No se encuentra el archivo: {path}')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'Listado' not in wb.sheetnames:
            raise CommandError('El Excel no contiene la hoja «Listado».')

        year, _ = AcademicYear.objects.update_or_create(
            name=options['year'],
            defaults={'is_current': True},
        )
        AcademicYear.objects.exclude(pk=year.pk).update(is_current=False)

        degrees = {}
        for study, (code, name) in STUDY_MAP.items():
            deg, _ = DegreeProgram.objects.update_or_create(
                code=code, defaults={'name': name, 'is_active': True},
            )
            degrees[study] = deg

        courses = {}
        for study, deg in degrees.items():
            deg_code = STUDY_MAP[study][0]
            for n in range(1, 5):
                shift = _course_shift(deg_code, n)
                course, _ = Course.objects.update_or_create(
                    degree_program=deg,
                    academic_year=year,
                    number=n,
                    defaults={'shift': shift},
                )
                courses[(study, n)] = course

        classrooms = {}
        for code, (room_code, room_name, cap) in DEGREE_CLASSROOMS.items():
            room, _ = Classroom.objects.update_or_create(
                code=room_code, defaults={'name': room_name, 'capacity': cap},
            )
            classrooms[code] = room

        if options['clear_eps']:
            deg_ids = list(
                DegreeProgram.objects.filter(code__in=EPS_DEGREE_CODES).values_list('pk', flat=True)
            )
            ScheduleEntry.objects.filter(
                schedule__academic_year=year,
                subject_offering__course__degree_program_id__in=deg_ids,
            ).delete()
            SubjectOffering.objects.filter(course__degree_program_id__in=deg_ids).delete()
            Subject.objects.filter(offerings__course__degree_program_id__in=deg_ids).distinct().delete()

        self._ensure_timeslots(year)

        ws = wb['Listado']
        prof_by_code = {}
        offering_count = 0
        subject_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[2] not in STUDY_MAP:
                continue
            study = row[2]
            course_num = _parse_course_number(row[3])
            if not course_num:
                continue
            subj_code = str(row[4]).strip()
            subj_name = str(row[5] or subj_code).strip()
            doble_val = row[10]
            h1, h2 = row[11], row[12]
            weekly = _weekly_sessions(h1, h2)
            group = _group_name(study, course_num, doble_val)

            raw_name = str(row[9] or '').strip()
            first, last, _ignored_email = _parse_professor_name(raw_name)
            code = _prof_code(raw_name)
            email = f'prof.{code.lower()}@ucjc.local' if code else _parse_professor_name(raw_name)[2]
            prof, _ = Professor.objects.update_or_create(
                email=email,
                defaults={
                    'first_name': first,
                    'last_name': last,
                    'display_name': raw_name[:160],
                },
            )
            if code:
                prof_by_code[code] = prof

            subject, created = Subject.objects.update_or_create(
                code=subj_code,
                defaults={
                    'name': subj_name[:120],
                    'weekly_sessions': weekly,
                    'session_duration_hours': 2,
                    'is_shared': False,
                },
            )
            if created:
                subject_count += 1
            if subject.weekly_sessions != weekly:
                subject.weekly_sessions = weekly
                subject.save(update_fields=['weekly_sessions', 'updated_at'])

            deg_code = STUDY_MAP[study][0]
            course = courses.get((study, course_num))
            if not course:
                continue

            room_key = f'{deg_code}-{course_num}-{_ascii_upper(group)[:12]}'
            classroom, _ = Classroom.objects.get_or_create(
                code=room_key,
                defaults={'name': f'Aula {room_key}', 'capacity': 30},
            )

            SubjectOffering.objects.update_or_create(
                subject=subject,
                course=course,
                group_name=group,
                defaults={
                    'professor': prof,
                    'classroom': classroom,
                },
            )
            offering_count += 1

        double_count = self._build_double_degree_offerings(year)
        offering_count += double_count

        schedule_entries = 0
        if options['apply_schedule']:
            schedule, _ = Schedule.objects.get_or_create(
                name=f'Horario EPS {options["year"]} (Excel Tomás)',
                academic_year=year,
                defaults={'status': 'DRAFT'},
            )
            ScheduleEntry.objects.filter(schedule=schedule).delete()
            schedule_entries = self._import_schedule_grids(wb, year, schedule, prof_by_code)

        self.stdout.write(self.style.SUCCESS(
            f'Importación completada · año {year.name} · '
            f'{subject_count} asignaturas nuevas · {offering_count} ofertas · '
            f'{ScheduleEntry.objects.filter(subject_offering__course__academic_year=year).count()} sesiones en BD'
            + (f' · {schedule_entries} celdas del Excel' if options['apply_schedule'] else '')
        ))
        self.stdout.write(
            'Titulaciones importadas: INF (Gininf), ROB (GRobotica), TEL (GTransporte), '
            f'DINFROB ({double_count} ofertas desde INF+ROB). '
            'Ejecuta con --apply-schedule para cargar la cuadrícula de referencia.'
        )

    def _build_double_degree_offerings(self, year):
        """
        El Excel no tiene filas «Doble Grado»: se compone con las ofertas de
        Informática y Robótica del mismo curso (1º–4º). 5º queda reservado (sin filas en Excel).
        """
        code, name = DOUBLE_DEGREE
        degree, _ = DegreeProgram.objects.update_or_create(
            code=code, defaults={'name': name, 'is_active': True},
        )
        dg_courses = {}
        for n in range(1, 6):
            shift = _course_shift(code, n)
            course, _ = Course.objects.update_or_create(
                degree_program=degree,
                academic_year=year,
                number=n,
                defaults={'shift': shift},
            )
            dg_courses[n] = course

        SubjectOffering.objects.filter(course__degree_program=degree, course__academic_year=year).delete()

        created = 0
        inf = DegreeProgram.objects.get(code='INF')
        rob = DegreeProgram.objects.get(code='ROB')
        for n in range(1, 5):
            dg_course = dg_courses[n]
            source_offerings = SubjectOffering.objects.filter(
                course__academic_year=year,
                course__number=n,
                course__degree_program__in=[inf, rob],
            ).select_related('subject', 'professor', 'classroom')
            for src in source_offerings:
                SubjectOffering.objects.update_or_create(
                    subject=src.subject,
                    course=dg_course,
                    group_name=src.group_name,
                    defaults={
                        'professor': src.professor,
                        'classroom': src.classroom,
                    },
                )
                created += 1
                if src.subject.name:
                    shared_qs = Subject.objects.filter(
                        offerings__course__degree_program__in=[inf, rob],
                        offerings__course__number=n,
                        name__iexact=src.subject.name,
                    ).distinct()
                    if shared_qs.count() >= 2 or (
                        SubjectOffering.objects.filter(
                            course__degree_program=inf, course__number=n, subject__name__iexact=src.subject.name,
                        ).exists()
                        and SubjectOffering.objects.filter(
                            course__degree_program=rob, course__number=n, subject__name__iexact=src.subject.name,
                        ).exists()
                    ):
                        src.subject.is_shared = True
                        src.subject.save(update_fields=['is_shared', 'updated_at'])
                        src.subject.degree_programs.set([inf, rob, degree])
        return created

    def _ensure_timeslots(self, year):
        ScheduleEntry.objects.filter(timeslot__academic_year=year).delete()
        ProfessorAvailability.objects.filter(timeslot__academic_year=year).delete()
        TimeSlotConfig.objects.filter(academic_year=year).delete()
        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        for day in days:
            for start, end in TIME_BLOCKS:
                TimeSlotConfig.objects.create(
                    academic_year=year,
                    day_of_week=day,
                    start_time=start,
                    end_time=end,
                    is_active=True,
                )

    def _import_schedule_grids(self, wb, year, schedule, prof_by_code):
        """Parsea hojas HORARIOS y crea ScheduleEntry."""
        slot_index = {}
        for slot in TimeSlotConfig.objects.filter(academic_year=year, is_active=True):
            slot_index[(slot.day_of_week, slot.start_time, slot.end_time)] = slot

        offering_index = {}
        for off in SubjectOffering.objects.filter(course__academic_year=year).select_related(
            'professor', 'course__degree_program', 'subject',
        ):
            pcode = _prof_code(f'{off.professor.last_name}, {off.professor.first_name}')
            key = (
                off.course.degree_program.code,
                off.course.number,
                off.group_name,
                pcode,
            )
            offering_index.setdefault(key, []).append(off)
            offering_index.setdefault((off.course.degree_program.code, off.course.number, off.group_name, ''), []).append(off)

        created = 0
        sheet_names = [n for n in wb.sheetnames if n.upper().startswith('HORARIOS')]
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            day_cols = {}
            header_row = None
            for i, row in enumerate(rows):
                if row and str(row[0] or '').upper() == 'HORA':
                    header_row = i
                    for ci, val in enumerate(row):
                        if ci >= 2 and val:
                            day_cols[ci] = DAY_MAP.get(_ascii_upper(str(val)).replace(' ', ''), None)
                    break
            if header_row is None:
                continue

            current_start = None
            for row in rows[header_row + 1:]:
                if not row:
                    continue
                if row[0] and hasattr(row[0], 'hour'):
                    current_start = row[0]
                group_label = str(row[1] or '').strip().upper()
                if not group_label or group_label == 'GRUPO' or group_label not in SCHEDULE_GROUP_MAP:
                    continue
                if not current_start:
                    continue
                deg_code, course_num, group_name = SCHEDULE_GROUP_MAP[group_label]
                block_end = self._block_end(current_start)
                for ci, day_code in day_cols.items():
                    if not day_code or ci >= len(row):
                        continue
                    cell = row[ci]
                    if not cell or not str(cell).strip():
                        continue
                    offering = self._match_offering(
                        str(cell), deg_code, course_num, group_name, prof_by_code, offering_index,
                    )
                    if not offering:
                        continue
                    slot = slot_index.get((day_code, current_start, block_end))
                    if not slot:
                        continue
                    try:
                        _, was_created = ScheduleEntry.objects.get_or_create(
                            schedule=schedule,
                            subject_offering=offering,
                            timeslot=slot,
                        )
                    except ValidationError:
                        continue
                    if was_created:
                        created += 1
        return created

    def _block_end(self, start):
        for s, e in TIME_BLOCKS:
            if s == start:
                return e
        h = start.hour + 2
        return time(h, start.minute)

    def _match_offering(self, cell_text, deg_code, course_num, group_name, prof_by_code, offering_index):
        tokens = str(cell_text).strip().split()
        prof_token = tokens[-1] if tokens else ''
        prof = prof_by_code.get(_ascii_upper(prof_token))
        pcode = _prof_code(f'{prof.last_name}, {prof.first_name}') if prof else _ascii_upper(prof_token)

        candidates = offering_index.get((deg_code, course_num, group_name, pcode), [])
        if not candidates and group_name.startswith('Grupo '):
            candidates = offering_index.get((deg_code, course_num, 'Grupo A', pcode), [])
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        hint = ' '.join(tokens[:-1]).lower()
        for off in candidates:
            if off.subject.name.lower()[:12] in hint or off.subject.code.lower() in hint:
                return off
        return candidates[0]
