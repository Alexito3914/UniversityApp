"""
Importación de catálogo docente desde Excel (Listado Tomás o «Listado 4 grados»).
Usado por el comando manage.py y por la subida web del decanato.
"""
import re
import unicodedata
from datetime import time
from io import BytesIO
from pathlib import Path

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction

from .models import (
    AcademicYear,
    Classroom,
    Course,
    DegreeProgram,
    Professor,
    ProfessorAvailability,
    Schedule,
    ScheduleEntry,
    Subject,
    SubjectOffering,
    TimeSlotConfig,
)
from .semesters import (
    EXCEL_SHEET_TITLES,
    SEMESTER_S1,
    SEMESTER_S2,
    semester_plans_from_horario_columns,
    weekly_sessions_from_horario_columns,
)
from .services import generate_schedule_entries

STUDY_MAP = {
    'Gininf': ('INF', 'Grado en Ingeniería Informática'),
    'GRobotica': ('ROB', 'Grado en Ingeniería Robótica'),
    'GTransporte': ('TEL', 'Grado en Ingeniería Telemática'),
    'DINFROB': ('DINFROB', 'Doble Grado Informática + Robótica'),
}

DOUBLE_DEGREE = ('DINFROB', 'Doble Grado Informática + Robótica')
EPS_DEGREE_CODES = ('INF', 'ROB', 'TEL', 'DINFROB')

DEGREE_CLASSROOMS = {
    'INF': ('LAB-INF', 'Laboratorio Informática', 30),
    'ROB': ('LAB-ROB', 'Laboratorio Robótica', 24),
    'TEL': ('LAB-TEL', 'Laboratorio Telemática', 28),
    'DINFROB': ('LAB-DG', 'Laboratorio Doble Grado', 30),
}

DAY_MAP = {
    'LUNES': 'MON',
    'MARTES': 'TUE',
    'MIÉRCOLES': 'WED',
    'MIERCOLES': 'WED',
    'JUEVES': 'THU',
    'VIERNES': 'FRI',
}

SCHEDULE_GROUP_MAP = {
    '1INF': ('INF', 1, 'Grupo A'),
    '2INF': ('INF', 2, 'Grupo A'),
    '3INF': ('INF', 3, 'Grupo A'),
    '4INF-TI': ('INF', 4, '4INF-TI'),
    '4INF-IS': ('INF', 4, '4INF-IS'),
    '1ROB': ('ROB', 1, 'Grupo 1'),
    '2ROB': ('ROB', 2, 'Grupo 2'),
    '3ROB': ('ROB', 3, 'Grupo 3'),
    '4ROB': ('ROB', 4, 'Grupo 4'),
}

TIME_BLOCKS = [
    (time(9, 0), time(11, 0)),
    (time(11, 0), time(13, 0)),
    (time(13, 0), time(15, 0)),
    (time(15, 30), time(17, 30)),
    (time(17, 30), time(19, 30)),
    (time(19, 30), time(21, 30)),
]

ORGANIZED_SHEET = 'Listado 4 grados'
LISTADO_SHEET = 'Listado'


class ExcelImportError(Exception):
    """Error de validación al importar un Excel."""


def _ascii_upper(text):
    if not text:
        return ''
    normalized = unicodedata.normalize('NFKD', str(text))
    return normalized.encode('ascii', 'ignore').decode('ascii').upper()


def _prof_code(full_name):
    if not full_name:
        return ''
    parts = str(full_name).split(',')
    if len(parts) < 2:
        return _ascii_upper(full_name).replace(' ', '')[:12]
    last = parts[0].strip().split()[0] if parts[0].strip() else ''
    first = parts[1].strip().split()[0] if parts[1].strip() else ''
    return (_ascii_upper(first[:1]) + _ascii_upper(last)) if last else ''


def _parse_course_number(raw):
    if not raw:
        return None
    m = re.search(r'(\d+)', str(raw))
    return int(m.group(1)) if m else None


def _course_shift(degree_code, course_num):
    if degree_code == 'DINFROB':
        return 'AFTERNOON' if course_num >= 4 else 'MORNING'
    return 'AFTERNOON' if course_num == 4 else 'MORNING'


def _group_name(study, course_num, doble_val):
    if study == 'Gininf' and course_num == 4:
        if doble_val == 4:
            return '4INF-TI'
        if doble_val == 5:
            return '4INF-IS'
    if study == 'GRobotica' and doble_val not in (None, True, False, ''):
        try:
            return f'Grupo {int(doble_val)}'
        except (TypeError, ValueError):
            pass
    if doble_val in (None, True, False, ''):
        return 'Grupo A'
    try:
        return f'Grupo {int(doble_val)}'
    except (TypeError, ValueError):
        return 'Grupo A'


def _parse_professor_name(raw):
    raw = str(raw or '').strip()
    if not raw:
        return 'Sin', 'Asignar', 'sin.asignar@ucjc.local'
    if ',' in raw:
        last_part, first_part = raw.split(',', 1)
        last_name = last_part.strip()
        first_name = first_part.strip().split()[0] if first_part.strip() else 'Docente'
    else:
        bits = raw.split()
        first_name = bits[0] if bits else 'Docente'
        last_name = ' '.join(bits[1:]) if len(bits) > 1 else bits[0]
    slug = _ascii_upper(last_name).replace(' ', '').replace(',', '')[:20].lower()
    email = f'{slug or "prof"}.{_ascii_upper(first_name[:1]).lower()}@ucjc.local'
    return first_name, last_name, email


def _organized_study_key(row):
    excel_study = str(row[9] or '').strip()
    if excel_study == 'DINFROB' or excel_study in STUDY_MAP:
        return excel_study
    grado = str(row[0] or '').lower()
    if 'doble' in grado:
        return 'DINFROB'
    if 'rob' in grado:
        return 'GRobotica'
    if 'telem' in grado or 'transporte' in grado:
        return 'GTransporte'
    if 'inform' in grado:
        return 'Gininf'
    return excel_study or None


def _normalize_listado_rows(wb):
    """Devuelve filas con índices del Listado Tomás (estudio=col 2, curso=3, etc.)."""
    if LISTADO_SHEET in wb.sheetnames:
        for row in wb[LISTADO_SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[2] not in STUDY_MAP:
                continue
            yield row
        return

    if ORGANIZED_SHEET in wb.sheetnames:
        for row in wb[ORGANIZED_SHEET].iter_rows(min_row=2, values_only=True):
            if not row or not row[3]:
                continue
            study = _organized_study_key(row)
            if study not in STUDY_MAP:
                continue
            course_num = _parse_course_number(row[1])
            if not course_num:
                continue
            yield (
                None,
                None,
                study,
                row[1],
                row[3],
                row[4],
                row[5],
                row[2],
                row[6],
                row[7],
                row[10],
                row[11],
                row[12],
            )
        return

    raise ExcelImportError(
        f'El Excel debe contener la hoja «{LISTADO_SHEET}» (formato Tomás) '
        f'o «{ORGANIZED_SHEET}» (formato organizado EPS).'
    )


class HorariosExcelImporter:
    def __init__(
        self,
        year_name,
        *,
        clear_eps=False,
        apply_schedule=False,
        auto_generate=False,
    ):
        self.year_name = year_name
        self.clear_eps = clear_eps
        self.apply_schedule = apply_schedule
        self.auto_generate = auto_generate

    @transaction.atomic
    def run(self, source):
        """
        Importa desde ruta, Path o file-like (.xlsx).
        Devuelve dict con estadísticas.
        """
        if isinstance(source, (str, Path)):
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        else:
            data = source.read()
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)

        listado_rows = list(_normalize_listado_rows(wb))
        if not listado_rows:
            raise ExcelImportError('No se encontraron filas de titulaciones EPS en el Excel.')

        year, _ = AcademicYear.objects.update_or_create(
            name=self.year_name,
            defaults={'is_current': True},
        )
        AcademicYear.objects.exclude(pk=year.pk).update(is_current=False)

        studies_in_file = {str(r[2]) for r in listado_rows}
        degrees = {}
        for study in studies_in_file:
            if study not in STUDY_MAP:
                continue
            code, name = STUDY_MAP[study]
            deg, _ = DegreeProgram.objects.update_or_create(
                code=code, defaults={'name': name, 'is_active': True},
            )
            degrees[study] = deg

        courses = {}
        for study, deg in degrees.items():
            deg_code = STUDY_MAP[study][0]
            max_n = 5 if deg_code == 'DINFROB' else 4
            for n in range(1, max_n + 1):
                shift = _course_shift(deg_code, n)
                course, _ = Course.objects.update_or_create(
                    degree_program=deg,
                    academic_year=year,
                    number=n,
                    defaults={'shift': shift},
                )
                courses[(study, n)] = course

        for code, (room_code, room_name, cap) in DEGREE_CLASSROOMS.items():
            Classroom.objects.get_or_create(
                code=room_code, defaults={'name': room_name, 'capacity': cap},
            )

        if self.clear_eps:
            deg_ids = list(
                DegreeProgram.objects.filter(code__in=EPS_DEGREE_CODES).values_list('pk', flat=True)
            )
            ScheduleEntry.objects.filter(
                schedule__academic_year=year,
                subject_offering__course__degree_program_id__in=deg_ids,
            ).delete()
            SubjectOffering.objects.filter(course__degree_program_id__in=deg_ids).delete()
            Subject.objects.filter(offerings__course__degree_program_id__in=deg_ids).distinct().delete()

        self._ensure_timeslots(year)

        prof_by_code = {}
        offering_count = 0
        subject_count = 0
        imported_degrees = set()

        for row in listado_rows:
            study = row[2]
            if study not in STUDY_MAP:
                continue
            course_num = _parse_course_number(row[3])
            if not course_num:
                continue
            subj_code = str(row[4]).strip()
            subj_name = str(row[5] or subj_code).strip()
            semester_col = row[7]
            doble_val = row[10]
            h1, h2 = row[11], row[12]
            group = _group_name(study, course_num, doble_val)

            raw_name = str(row[9] or '').strip()
            first, last, _ignored = _parse_professor_name(raw_name)
            pcode = _prof_code(raw_name)
            email = f'prof.{pcode.lower()}@ucjc.local' if pcode else _parse_professor_name(raw_name)[2]
            prof, _ = Professor.objects.update_or_create(
                email=email,
                defaults={
                    'first_name': first,
                    'last_name': last,
                    'display_name': raw_name[:160],
                },
            )
            if pcode:
                prof_by_code[pcode] = prof

            weekly = weekly_sessions_from_horario_columns(h1, h2)
            subject, created = Subject.objects.update_or_create(
                code=subj_code,
                defaults={
                    'name': subj_name[:120],
                    'weekly_sessions': weekly,
                    'session_duration_hours': 2,
                    'is_shared': False,
                },
            )
            if created:
                subject_count += 1
            if subject.weekly_sessions != weekly:
                subject.weekly_sessions = weekly
                subject.save(update_fields=['weekly_sessions', 'updated_at'])

            deg = degrees.get(study)
            if deg:
                subject.degree_programs.add(deg)

            deg_code = STUDY_MAP[study][0]
            course = courses.get((study, course_num))
            if not course:
                continue

            room_key = f'{deg_code}-{course_num}-{_ascii_upper(group)[:12]}'
            classroom, _ = Classroom.objects.get_or_create(
                code=room_key,
                defaults={'name': f'Aula {room_key}', 'capacity': 30},
            )

            for semester, weekly_sem in semester_plans_from_horario_columns(
                h1, h2, semester_col=semester_col,
            ):
                SubjectOffering.objects.update_or_create(
                    subject=subject,
                    course=course,
                    group_name=group,
                    semester=semester,
                    defaults={
                        'professor': prof,
                        'classroom': classroom,
                        'weekly_sessions': weekly_sem,
                    },
                )
                offering_count += 1
                imported_degrees.add(deg_code)

        double_count = 0
        if 'DINFROB' not in studies_in_file:
            double_count = self._build_double_degree_offerings(year)

        schedule_entries = 0
        schedules_created = {}
        if self.apply_schedule:
            for semester in (SEMESTER_S1, SEMESTER_S2):
                label = 'Primer' if semester == SEMESTER_S1 else 'Segundo'
                schedule, _ = Schedule.objects.update_or_create(
                    name=f'Horario EPS {self.year_name} — {label} cuatrimestre',
                    academic_year=year,
                    semester=semester,
                    defaults={'status': 'DRAFT'},
                )
                ScheduleEntry.objects.filter(schedule=schedule).delete()
                schedule_entries += self._import_schedule_grids(
                    wb, year, schedule, prof_by_code, semester=semester,
                )
                schedules_created[semester] = schedule.pk

        generated = {}
        if self.auto_generate:
            for semester in (SEMESTER_S1, SEMESTER_S2):
                label = 'Primer' if semester == SEMESTER_S1 else 'Segundo'
                schedule, _ = Schedule.objects.update_or_create(
                    name=f'Horario EPS {self.year_name} — {label} cuatrimestre',
                    academic_year=year,
                    semester=semester,
                    defaults={'status': 'DRAFT'},
                )
                if self.auto_generate and not self.apply_schedule:
                    ScheduleEntry.objects.filter(schedule=schedule).delete()
                result = generate_schedule_entries(schedule, clear_existing=True)
                generated[semester] = {
                    'schedule_id': schedule.pk,
                    'created': result['created'],
                    'unresolved': len(result['unresolved']),
                }

        return {
            'year': year.name,
            'subjects_new': subject_count,
            'offerings': offering_count + double_count,
            'double_degree_offerings': double_count,
            'degrees': sorted(imported_degrees),
            'schedule_grid_cells': schedule_entries,
            'generated': generated,
            'rows_processed': len(listado_rows),
        }

    def _build_double_degree_offerings(self, year):
        code, name = DOUBLE_DEGREE
        degree, _ = DegreeProgram.objects.update_or_create(
            code=code, defaults={'name': name, 'is_active': True},
        )
        dg_courses = {}
        for n in range(1, 6):
            shift = _course_shift(code, n)
            course, _ = Course.objects.update_or_create(
                degree_program=degree,
                academic_year=year,
                number=n,
                defaults={'shift': shift},
            )
            dg_courses[n] = course

        SubjectOffering.objects.filter(course__degree_program=degree, course__academic_year=year).delete()

        created = 0
        inf = DegreeProgram.objects.get(code='INF')
        rob = DegreeProgram.objects.get(code='ROB')
        for n in range(1, 5):
            dg_course = dg_courses[n]
            source_offerings = SubjectOffering.objects.filter(
                course__academic_year=year,
                course__number=n,
                course__degree_program__in=[inf, rob],
            ).select_related('subject', 'professor', 'classroom')
            for src in source_offerings:
                SubjectOffering.objects.update_or_create(
                    subject=src.subject,
                    course=dg_course,
                    group_name=src.group_name,
                    semester=src.semester,
                    defaults={
                        'professor': src.professor,
                        'classroom': src.classroom,
                        'weekly_sessions': src.sessions_per_week(),
                    },
                )
                created += 1
        return created

    def _ensure_timeslots(self, year):
        if TimeSlotConfig.objects.filter(academic_year=year, is_active=True).exists():
            return
        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        for day in days:
            for start, end in TIME_BLOCKS:
                TimeSlotConfig.objects.create(
                    academic_year=year,
                    day_of_week=day,
                    start_time=start,
                    end_time=end,
                    is_active=True,
                )

    def _import_schedule_grids(self, wb, year, schedule, prof_by_code, semester=SEMESTER_S1):
        slot_index = {}
        for slot in TimeSlotConfig.objects.filter(academic_year=year, is_active=True):
            slot_index[(slot.day_of_week, slot.start_time, slot.end_time)] = slot

        offering_index = {}
        for off in SubjectOffering.objects.filter(
            course__academic_year=year, semester=semester,
        ).select_related('professor', 'course__degree_program', 'subject'):
            pcode = _prof_code(f'{off.professor.last_name}, {off.professor.first_name}')
            key = (off.course.degree_program.code, off.course.number, off.group_name, pcode)
            offering_index.setdefault(key, []).append(off)
            offering_index.setdefault(
                (off.course.degree_program.code, off.course.number, off.group_name, ''), []
            ).append(off)

        created = 0
        target_sheet = EXCEL_SHEET_TITLES.get(semester)
        if target_sheet and target_sheet in wb.sheetnames:
            sheet_names = [target_sheet]
        else:
            sheet_names = [n for n in wb.sheetnames if n.upper().startswith('HORARIOS')]
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            day_cols = {}
            header_row = None
            for i, row in enumerate(rows):
                if row and str(row[0] or '').upper() == 'HORA':
                    header_row = i
                    for ci, val in enumerate(row):
                        if ci >= 2 and val:
                            day_cols[ci] = DAY_MAP.get(_ascii_upper(str(val)).replace(' ', ''), None)
                    break
            if header_row is None:
                continue

            current_start = None
            for row in rows[header_row + 1:]:
                if not row:
                    continue
                if row[0] and hasattr(row[0], 'hour'):
                    current_start = row[0]
                group_label = str(row[1] or '').strip().upper()
                if not group_label or group_label == 'GRUPO' or group_label not in SCHEDULE_GROUP_MAP:
                    continue
                if not current_start:
                    continue
                deg_code, course_num, group_name = SCHEDULE_GROUP_MAP[group_label]
                block_end = self._block_end(current_start)
                for ci, day_code in day_cols.items():
                    if not day_code or ci >= len(row):
                        continue
                    cell = row[ci]
                    if not cell or not str(cell).strip():
                        continue
                    offering = self._match_offering(
                        str(cell), deg_code, course_num, group_name, prof_by_code, offering_index,
                    )
                    if not offering:
                        continue
                    slot = slot_index.get((day_code, current_start, block_end))
                    if not slot:
                        continue
                    try:
                        _, was_created = ScheduleEntry.objects.get_or_create(
                            schedule=schedule,
                            subject_offering=offering,
                            timeslot=slot,
                        )
                    except ValidationError:
                        continue
                    if was_created:
                        created += 1
        return created

    def _block_end(self, start):
        for s, e in TIME_BLOCKS:
            if s == start:
                return e
        h = start.hour + 2
        return time(h, start.minute)

    def _match_offering(self, cell_text, deg_code, course_num, group_name, prof_by_code, offering_index):
        tokens = str(cell_text).strip().split()
        prof_token = tokens[-1] if tokens else ''
        prof = prof_by_code.get(_ascii_upper(prof_token))
        pcode = _prof_code(f'{prof.last_name}, {prof.first_name}') if prof else _ascii_upper(prof_token)

        candidates = offering_index.get((deg_code, course_num, group_name, pcode), [])
        if not candidates and group_name.startswith('Grupo '):
            candidates = offering_index.get((deg_code, course_num, 'Grupo A', pcode), [])
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        hint = ' '.join(tokens[:-1]).lower()
        for off in candidates:
            if off.subject.name.lower()[:12] in hint or off.subject.code.lower() in hint:
                return off
        return candidates[0]
