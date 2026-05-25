"""Utilidades compartidas para import/export de cuadrículas Excel (formato decanato)."""
import re
import unicodedata
from datetime import time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import ScheduleEntry
from .models import Schedule as ScheduleModel
from .semesters import EXCEL_SHEET_TITLES, SEMESTER_LABELS, SEMESTER_S1, SEMESTER_S2

DAY_MAP = {
    'LUNES': 'MON',
    'MARTES': 'TUE',
    'MIÉRCOLES': 'WED',
    'MIERCOLES': 'WED',
    'JUEVES': 'THU',
    'VIERNES': 'FRI',
}

DAY_HEADERS = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES']
DAY_CODES = [DAY_MAP[day] for day in DAY_HEADERS]

SCHEDULE_GROUP_MAP = {
    '1INF': ('INF', 1, 'Grupo A'),
    '2INF': ('INF', 2, 'Grupo A'),
    '3INF': ('INF', 3, 'Grupo A'),
    '4INF-TI': ('INF', 4, '4INF-TI'),
    '4INF-IS': ('INF', 4, '4INF-IS'),
    '1ROB': ('ROB', 1, 'Grupo 1'),
    '2ROB': ('ROB', 2, 'Grupo 2'),
    '3ROB': ('ROB', 3, 'Grupo 3'),
    '4ROB': ('ROB', 4, 'Grupo 4'),
}

OFFERING_TO_GROUP_LABEL = {
    (deg, num, group): label for label, (deg, num, group) in SCHEDULE_GROUP_MAP.items()
}

TIME_BLOCKS = [
    (time(9, 0), time(11, 0)),
    (time(11, 0), time(13, 0)),
    (time(13, 0), time(15, 0)),
    (time(15, 30), time(17, 30)),
    (time(17, 30), time(19, 30)),
    (time(19, 30), time(21, 30)),
]

MORNING_GROUP_ORDER = [
    '1INF', '2INF', '3INF', '1ROB', '2ROB', '3ROB',
    '1TEL', '2TEL', '3TEL', '1DINFROB', '2DINFROB', '3DINFROB',
]
AFTERNOON_GROUP_ORDER = [
    '4INF-TI', '4INF-IS', '4ROB', '4TEL', '5DINFROB', '4DINFROB',
]

DEGREE_FILLS = {
    'INF': '3B82F6',
    'ROB': '10B981',
    'TEL': 'F59E0B',
    'DINFROB': '8B5CF6',
}


def ascii_upper(text):
    if not text:
        return ''
    normalized = unicodedata.normalize('NFKD', str(text))
    return normalized.encode('ascii', 'ignore').decode('ascii').upper()


def prof_code(professor):
    if professor.display_name:
        raw = professor.display_name
    else:
        raw = f'{professor.last_name}, {professor.first_name}'
    if not raw:
        return ''
    parts = str(raw).split(',')
    if len(parts) < 2:
        return ascii_upper(raw).replace(' ', '')[:12]
    last = parts[0].strip().split()[0] if parts[0].strip() else ''
    first = parts[1].strip().split()[0] if parts[1].strip() else ''
    return (ascii_upper(first[:1]) + ascii_upper(last)) if last else ''


def short_subject_name(name, max_len=22):
    text = re.sub(r'\s+', ' ', str(name or '').strip())
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + '…'


def cell_text_for_entry(entry):
    offering = entry.subject_offering
    pcode = prof_code(offering.professor)
    short = short_subject_name(offering.subject.name)
    return f'{short} {pcode}'.strip()


def group_label_for_offering(offering):
    deg = offering.course.degree_program.code
    num = offering.course.number
    group = offering.group_name
    key = (deg, num, group)
    if key in OFFERING_TO_GROUP_LABEL:
        return OFFERING_TO_GROUP_LABEL[key]
    if deg == 'ROB' and group.startswith('Grupo '):
        try:
            n = int(group.split()[-1])
            return f'{n}ROB'
        except ValueError:
            pass
    if group in ('Grupo A', 'Grupo 1') and deg in ('TEL', 'DINFROB'):
        return f'{num}{deg}'
    safe_group = ascii_upper(group).replace(' ', '')[:10]
    return f'{num}{deg}-{safe_group}' if safe_group else f'{num}{deg}'


def _entries_queryset(schedule, degree_id=None, course_id=None):
    qs = ScheduleEntry.objects.filter(schedule=schedule).select_related(
        'timeslot',
        'subject_offering__subject',
        'subject_offering__professor',
        'subject_offering__course__degree_program',
    )
    if degree_id:
        qs = qs.filter(subject_offering__course__degree_program_id=degree_id)
    if course_id:
        qs = qs.filter(subject_offering__course_id=course_id)
    return qs.order_by('timeslot__start_time', 'timeslot__day_of_week')


def build_grid_cells(schedule, degree_id=None, course_id=None):
    """Devuelve {group_label: {(start, end): {day_code: cell_text}}}."""
    grid = {}
    for entry in _entries_queryset(schedule, degree_id, course_id):
        label = group_label_for_offering(entry.subject_offering)
        slot = entry.timeslot
        key = (slot.start_time, slot.end_time)
        grid.setdefault(label, {}).setdefault(key, {})[slot.day_of_week] = cell_text_for_entry(entry)
    return grid


def _ordered_groups(grid, morning):
    order = MORNING_GROUP_ORDER if morning else AFTERNOON_GROUP_ORDER
    present = [g for g in order if g in grid]
    extras = sorted(g for g in grid if g not in order)
    return present + extras


def _degree_fill_for_group(label):
    if 'DINFROB' in label:
        return DEGREE_FILLS['DINFROB']
    if 'TEL' in label:
        return DEGREE_FILLS['TEL']
    if 'INF' in label:
        return DEGREE_FILLS['INF']
    if 'ROB' in label:
        return DEGREE_FILLS['ROB']
    return 'E5E7EB'


def _write_grid_sheet(ws, schedule, grid, semester_label):
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='F3F4F6')
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=f'HORARIOS {schedule.academic_year.name} — {schedule.name}').font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.cell(row=2, column=1, value=semester_label).font = Font(bold=True, size=11)

    header_row = 4
    ws.cell(row=header_row, column=1, value='HORA').font = header_font
    ws.cell(row=header_row, column=2, value='GRUPO').font = header_font
    for ci, day in enumerate(DAY_HEADERS, start=3):
        cell = ws.cell(row=header_row, column=ci, value=day)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row_idx = header_row + 1
    morning_groups = _ordered_groups(grid, morning=True)
    afternoon_groups = _ordered_groups(grid, morning=False)

    for start, end in TIME_BLOCKS:
        morning = start.hour < 15
        groups = morning_groups if morning else afternoon_groups
        if not groups:
            continue
        block_started = False
        for group_label in groups:
            row_cells = grid.get(group_label, {}).get((start, end), {})
            ws.cell(row=row_idx, column=1, value=start if not block_started else None)
            ws.cell(row=row_idx, column=2, value=group_label).font = Font(bold=True, size=10)
            fill = PatternFill('solid', fgColor=_degree_fill_for_group(group_label))
            for di, day_code in enumerate(DAY_CODES):
                cell = ws.cell(row=row_idx, column=di + 3, value=row_cells.get(day_code, ''))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border
                if row_cells.get(day_code):
                    cell.fill = fill
            row_idx += 1
            block_started = True
        row_idx += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    for col in ('C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[col].width = 28
    ws.freeze_panes = 'C5'


def build_schedule_workbook(schedule, degree_id=None, course_id=None):
    year = schedule.academic_year
    by_semester = {
        s.semester: s
        for s in ScheduleModel.objects.filter(academic_year=year)
    }
    s1 = by_semester.get(SEMESTER_S1, schedule if schedule.semester == SEMESTER_S1 else None)
    s2 = by_semester.get(SEMESTER_S2, schedule if schedule.semester == SEMESTER_S2 else None)
    if not s1 and schedule.semester == SEMESTER_S1:
        s1 = schedule
    if not s2 and schedule.semester == SEMESTER_S2:
        s2 = schedule

    wb = Workbook()
    ws1 = wb.active
    ws1.title = EXCEL_SHEET_TITLES[SEMESTER_S1]
    grid_s1 = build_grid_cells(s1, degree_id=degree_id, course_id=course_id) if s1 else {}
    _write_grid_sheet(ws1, schedule, grid_s1, f'{SEMESTER_LABELS[SEMESTER_S1]} (exportado desde UCJC Horarios)')

    ws2 = wb.create_sheet(EXCEL_SHEET_TITLES[SEMESTER_S2])
    grid_s2 = build_grid_cells(s2, degree_id=degree_id, course_id=course_id) if s2 else {}
    _write_grid_sheet(ws2, schedule, grid_s2, f'{SEMESTER_LABELS[SEMESTER_S2]} (exportado desde UCJC Horarios)')
    return wb


def workbook_to_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
