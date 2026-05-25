import json

from django.test import TestCase
from django.urls import reverse

from horarios.models import Notification, Schedule, ScheduleEntry, TimeSlotConfig

from .base import add_schedule_entry, build_schedule_lab, create_user


class ScheduleManagementViewTests(TestCase):
    def setUp(self):
        self.lab = build_schedule_lab()
        self.schedule = self.lab['schedule']
        self.dean = create_user('dean_sm', 'DEAN')
        self.it = create_user('it_sm', 'IT')
        self.entry = add_schedule_entry(self.schedule, self.lab['offering_a'], self.lab['morning_slot'])

    def test_schedule_detail_shows_entries_for_dean(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.get(reverse('horarios:schedule_detail', kwargs={'pk': self.schedule.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['can_manage'])
        self.assertContains(response, 'TST-A')

    def test_schedule_generate_creates_entries(self):
        self.client.login(username='it_sm', password='ucjc1234')
        response = self.client.post(reverse('horarios:schedule_generate', kwargs={'pk': self.schedule.pk}), {
            'clear_existing': 'on',
            'academic_year': self.lab['year'].pk,
        })
        self.assertRedirects(response, reverse('horarios:schedule_detail', kwargs={'pk': self.schedule.pk}))
        self.assertGreater(self.schedule.entries.count(), 0)

    def test_schedule_entry_delete(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.post(reverse(
            'horarios:schedule_entry_delete',
            kwargs={'pk': self.schedule.pk, 'entry_pk': self.entry.pk},
        ))
        self.assertRedirects(response, reverse('horarios:schedule_detail', kwargs={'pk': self.schedule.pk}))
        self.assertFalse(ScheduleEntry.objects.filter(pk=self.entry.pk).exists())

    def test_schedule_entry_move(self):
        self.client.login(username='it_sm', password='ucjc1234')
        response = self.client.post(
            reverse('horarios:schedule_entry_move', kwargs={'pk': self.schedule.pk, 'entry_pk': self.entry.pk}),
            data=json.dumps({'day': 'TUE', 'start': '09:00', 'end': '11:00'}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.timeslot.day_of_week, 'TUE')

    def test_schedule_export_csv(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.get(reverse('horarios:schedule_export_csv', kwargs={'pk': self.schedule.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')
        content = response.content.decode('utf-8-sig')
        self.assertIn('Test A', content)
        self.assertIn('INF-T', content)

    def test_schedule_export_excel_grid(self):
        import openpyxl
        from io import BytesIO

        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.get(reverse('horarios:schedule_export_excel', kwargs={'pk': self.schedule.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertIn('HORARIOS 1º CUATRIMESTRE', wb.sheetnames)
        ws = wb['HORARIOS 1º CUATRIMESTRE']
        flat = ' '.join(
            str(v) for row in ws.iter_rows(min_row=1, max_row=40, values_only=True) for v in row if v
        )
        self.assertIn('GRUPO', flat)
        self.assertIn('Test A', flat)
        self.assertIn('LUNES', flat)

    def test_schedule_api_json(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.get(reverse('horarios:schedule_api_json', kwargs={'pk': self.schedule.pk}))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['id'], self.schedule.pk)
        self.assertEqual(len(data['entries']), 1)

    def test_schedule_update_only_dean(self):
        self.client.login(username='it_sm', password='ucjc1234')
        response = self.client.get(reverse('horarios:schedule_update', kwargs={'pk': self.schedule.pk}))
        self.assertRedirects(response, reverse('horarios:schedule_list'))

    def test_schedule_delete_only_dean(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        pk = self.schedule.pk
        response = self.client.post(reverse('horarios:schedule_delete', kwargs={'pk': pk}))
        self.assertRedirects(response, reverse('horarios:schedule_list'))
        self.assertFalse(Schedule.objects.filter(pk=pk).exists())

    def test_dean_can_update_schedule_name(self):
        self.client.login(username='dean_sm', password='ucjc1234')
        response = self.client.post(reverse('horarios:schedule_update', kwargs={'pk': self.schedule.pk}), {
            'name': 'Horario renombrado',
            'academic_year': self.lab['year'].pk,
            'semester': 'S1',
        })
        self.assertRedirects(response, reverse('horarios:schedule_list'))
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.name, 'Horario renombrado')


class NotificationViewTests(TestCase):
    def setUp(self):
        self.user = create_user('notif_user', 'DEAN')
        self.notification = Notification.objects.create(
            user=self.user, message='Horario pendiente', notif_type='warn',
        )

    def test_notifications_list(self):
        self.client.login(username='notif_user', password='ucjc1234')
        response = self.client.get(reverse('horarios:notifications'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Horario pendiente')

    def test_notification_toggle(self):
        self.client.login(username='notif_user', password='ucjc1234')
        self.client.get(reverse('horarios:notification_toggle', kwargs={'pk': self.notification.pk}))
        self.notification.refresh_from_db()
        self.assertTrue(self.notification.is_read)

    def test_notification_mark_all(self):
        self.client.login(username='notif_user', password='ucjc1234')
        response = self.client.post(reverse('horarios:notification_mark_all'))
        self.assertRedirects(response, reverse('horarios:notifications'))
        self.assertFalse(Notification.objects.filter(user=self.user, is_read=False).exists())


class ConfigurationViewTests(TestCase):
    def setUp(self):
        self.lab = build_schedule_lab()
        create_user('dean_cfg', 'DEAN')
        create_user('prof_cfg', 'PROF', email=self.lab['prof_1'].email)

    def test_dean_can_add_timeslot(self):
        self.client.login(username='dean_cfg', password='ucjc1234')
        before = TimeSlotConfig.objects.filter(academic_year=self.lab['year']).count()
        response = self.client.post(reverse('horarios:configuration'), {
            'action': 'add',
            'day_of_week': 'WED',
            'start_time': '11:00',
            'end_time': '13:00',
        })
        self.assertRedirects(response, reverse('horarios:configuration'))
        self.assertEqual(TimeSlotConfig.objects.filter(academic_year=self.lab['year']).count(), before + 1)

    def test_professor_denied_configuration(self):
        self.client.login(username='prof_cfg', password='ucjc1234')
        response = self.client.get(reverse('horarios:configuration'))
        self.assertRedirects(response, reverse('horarios:home'))

    def test_dean_can_toggle_timeslot(self):
        slot = self.lab['morning_slot']
        self.client.login(username='dean_cfg', password='ucjc1234')
        self.client.post(reverse('horarios:configuration'), {'action': 'toggle', 'ts_id': slot.pk})
        slot.refresh_from_db()
        self.assertFalse(slot.is_active)


class MiscViewTests(TestCase):
    def setUp(self):
        self.lab = build_schedule_lab()
        create_user('dean_misc', 'DEAN')

    def test_degree_list(self):
        self.client.login(username='dean_misc', password='ucjc1234')
        response = self.client.get(reverse('horarios:degree_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'INF-T')

    def test_reports_view(self):
        self.client.login(username='dean_misc', password='ucjc1234')
        response = self.client.get(reverse('horarios:reports'))
        self.assertEqual(response.status_code, 200)

    def test_home_dashboard_metrics(self):
        self.client.login(username='dean_misc', password='ucjc1234')
        response = self.client.get(reverse('horarios:home'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Resumen')

    def test_subject_offering_create_get(self):
        self.client.login(username='dean_misc', password='ucjc1234')
        response = self.client.get(reverse('horarios:subject_offering_create'))
        self.assertEqual(response.status_code, 200)
